import json
import os
from openpyxl import load_workbook, Workbook
import unicodedata
import re
from copy import copy
from openpyxl.cell.cell import Cell
from datetime import datetime, timedelta
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.utils import get_column_letter
from collections import defaultdict

import unicodedata

def _normalize_name(name):
    # Convert full-width to half-width, strip, remove extra spaces
    name = unicodedata.normalize("NFKC", name)
    name = name.replace("　", " ")  # full-width space to half-width
    name = " ".join(name.strip().split())  # remove double spaces
    return name

TIME_ROW_MAP = {
    "13:10":5 ,
    "14:40":6,
    "16:30":7,
    "18:00":8,
    "19:30":9,
}

def copy_worksheet_template(target_wb, template_ws, new_title):
    new_ws = target_wb.create_sheet(title=new_title)
    merged_ranges = set()
    for merged_range in template_ws.merged_cells.ranges:
        merged_ranges.update(merged_range.cells)
    for row in template_ws.iter_rows():
        for cell in row:
            if not isinstance(cell, Cell):
                continue
            is_merged = (cell.coordinate in merged_ranges)
            new_cell = new_ws.cell(row=cell.row, column=cell.column)
            if not is_merged:
                new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for merged_range in template_ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_range))
    return new_ws

def get_top_left_if_merged(ws, row, col):
    cell_coord = f"{get_column_letter(col)}{row}"
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws.cell(row=row, column=col)



class Schedule_result:
    def __init__(self, student_data, teacher_data, match_data, student_template, teacher_template, date_list):
        # Ensure that all names in student_data are normalized properly.
        self.student_data = {
            _normalize_name(k): v for k, v in student_data.items()
        }
        self.teacher_data = {
            _normalize_name(k): v for k, v in teacher_data.items()
        }
        self.subject_data = match_data
        self.student_template = student_template
        self.teacher_template = teacher_template
        self.date_list = date_list
        self.teacher_output_path = "output/teachers_schedule.xlsx"
        self.student_output_dirs = {
            'elementary': "output/students_elementary.xlsx",
            'middle': "output/students_middle.xlsx",
            'high': "output/students_high.xlsx",
        }
        self.schedule_data = []
        self.date_order = []
        self.used_teacher_slots = {}  # (teacher, date, time, booth_index) -> True

    def run(self):
        # Normalize student names before running the schedule
        self.normalize_student_names()
        self.generate_schedule()
        self.generate_teacher_excel()
        self.generate_student_excels()

    def is_slot_available(self, student, teacher, date, time, booth_index):
        student = _normalize_name(student)
        teacher = _normalize_name(teacher)
        normalized_time = unicodedata.normalize("NFKC", time)

        available_times_student = list(self.student_data.get(student, {}).get("schedule", {}).get(date, {}).keys())
        available_times_teacher = list(self.teacher_data.get(teacher, {}).get("schedule", {}).get(date, {}).keys())
        
        student_avail = self.student_data.get(student, {}).get("schedule", {}).get(date, {}).get(time, False)
        teacher_avail = self.teacher_data.get(teacher, {}).get("schedule", {}).get(date, {}).get(time, False)
        
        # Fix: Ensure teacher availability at the booth
        if booth_index >= len(teacher_avail):
            # print(f"⚠️ Booth {booth_index} out of range for {teacher} on {date} at {time}.")
            return False
            
        return student_avail and teacher_avail[booth_index]

    def normalize_student_names(self):
        # print("📋 Normalizing student names...")
        # Step 1: Normalize keys in student_data
        normalized_data = {}
        for k in self.student_data:
            norm_k = _normalize_name(k)
            normalized_data[norm_k] = self.student_data[k]
            # print(f" - {k} ➜ {norm_k}")
        self.student_data = normalized_data  # Replace with normalized keys
        # Step 2: Normalize student names inside subject_data
        for entry in self.subject_data:
            entry['student_name'] = _normalize_name(entry['student_name'])

    def generate_schedule(self):
        # 1. Group all subject-teacher pairs by student
        student_subject_map = defaultdict(list)
        for entry in self.subject_data:
            student = entry['student_name']
            grade = entry.get('grade', 'other')
            for subj in entry.get('subjects', []):
                teacher_raw = subj.get('teacher')
                teacher = _normalize_name(teacher_raw.get('name') if isinstance(teacher_raw, dict) else teacher_raw)
                subject_name = subj.get('name')
                subject_type = '特別' if subj.get('special_classes', 0) > 0 else '通常'
                count = subj.get('regular_classes', 0) + subj.get('special_classes', 0)
                if count > 0:
                    student_subject_map[student].append({
                        'teacher': teacher,
                        'subject': subject_name,
                        'type': subject_type,
                        'count': count,
                        'grade': grade
                    })
        # 2. Group by teacher -> students list
        teacher_to_subjects = defaultdict(list)
        for student, subj_list in student_subject_map.items():
            for subj in subj_list:
                teacher = subj['teacher']
                teacher_to_subjects[teacher].append((student, subj))

        # 3. Sort teachers by number of total lectures needed
        teacher_priority = sorted(
            teacher_to_subjects.items(),
            key=lambda x: -sum(sub['count'] for _, sub in x[1])
        )

        schedule_map = []  # final list of dicts like self.schedule_data
        teacher_slot_tracker = {}  # (teacher, date, time) -> list of students

        def assign(student, teacher, date, time, booth_index, subject, subject_type, grade):
            print(student, teacher, date, time, booth_index, subject, subject_type, grade)
            self.student_data[student]['schedule'][date][time] = False
            self.teacher_data[teacher]['schedule'][date][time][booth_index] = False
            key = (teacher, date, time, booth_index)
            self.used_teacher_slots[key] = True
            teacher_slot_tracker.setdefault((teacher, date, time), []).append(student)
            schedule_map.append({
                'date': date,
                'time': time,
                'student': student,
                'teacher': teacher,
                'subject': subject,
                'type': subject_type,
                'grade': grade
            })

        for teacher, student_subjects in teacher_priority:
            teacher = _normalize_name(teacher)
            student_subjects.sort(key=lambda x: x[1]['count'])

            for student, subject_data in student_subjects:
                student = _normalize_name(student)
                remaining = subject_data['count']
                subject = subject_data['subject']
                subject_type = subject_data['type']
                grade = subject_data['grade']
                last_scheduled_dates = []
                if student not in self.student_data:
                    # print(f"⚠️ Student not found in student_data: {student}")
                    continue

                for date in sorted(set(self.date_list) & set(self.student_data[student]['schedule'].keys())):
                    if remaining <= 0:
                        break
                    if date not in self.teacher_data[teacher]['schedule']:
                        continue

                    for time in TIME_ROW_MAP:
                        normalized_time = unicodedata.normalize("NFKC", time)
                        if normalized_time not in self.teacher_data[teacher]['schedule'][date]:
                            # print(f"❌ Time '{normalized_time}' not found on {date} for {teacher}")/
                            continue

                        if remaining <= 0:
                            break
                        for booth_index in range(len(self.teacher_data[teacher]['schedule'][date][normalized_time])):
                            key = (teacher, date, time)
                            print(len(teacher_slot_tracker.get(key, [])) >= 2)
                            if len(teacher_slot_tracker.get(key, [])) >= 2:
                                print(f"🛑 Skipped: {teacher} has 2 students already at {date} {time}")
                                continue

                            date_obj = datetime.strptime(date, "%Y-%m-%d")
                            if any(abs((date_obj - prev).days) < 2 for prev in last_scheduled_dates):
                                # print(f"⏭️ Skipped: {student}'s lesson on {date} too close to previous")
                                continue
                            if subject_data['count'] > 12 and any(abs((date_obj - prev).days) < 4 for prev in last_scheduled_dates):
                                # print(f"⏭️ Skipped: {student}'s >12 lessons, too close to previous")
                                continue
                           
                            if self.is_slot_available(student, teacher, date, normalized_time, booth_index):
                                print(f"❌ No slot: {student} or {teacher} not available on {date} {time} booth {booth_index}")
                                assign(student, teacher, date, time, booth_index, subject, subject_type, grade)
                                last_scheduled_dates.append(date_obj)
                                remaining -= 1
                                break

                if remaining > 0:
                    print(f"⚠️ Could not fully schedule {student} for {subject} with {teacher}, {remaining} left.")

        self.schedule_data = schedule_map
        self.date_order = sorted({entry['date'] for entry in schedule_map})

    def generate_teacher_excel(self):
        if not self.schedule_data:
            print("⚠️ No teacher data found — no file saved.")
            return
        if isinstance(self.teacher_template, str):
            self.teacher_template = load_workbook(self.teacher_template)
        template_wb = self.teacher_template
        template_sheetnames = template_wb.sheetnames
        name_map = {}

        for entry in self.schedule_data:
            full_name = _normalize_name(entry['teacher'])
            match = None
            for sheet_name in template_sheetnames:
                normalized_sheet = _normalize_name(sheet_name)
                if full_name == normalized_sheet or normalized_sheet in full_name:
                    match = sheet_name
                    full_name = _normalize_name(entry['teacher'])  
                    break

            if match:
                name_map[full_name] = match
            else:
                print(f"⚠️ No matching sheet found for teacher: {full_name}")
        output_wb = Workbook()
        output_wb.remove(output_wb.active)
        handled_teachers = set()
        for entry in self.schedule_data:
            teacher = entry['teacher']
            if teacher not in name_map:
                continue
            template_sheetname = name_map[teacher]
            if teacher not in handled_teachers:
                template_ws = self.teacher_template[template_sheetname]
                new_ws = copy_worksheet_template(output_wb, template_ws, teacher[:30])
                handled_teachers.add(teacher)
            else:
                # Always access the sheet by name, not by previously cached variable
                new_ws = output_wb[teacher[:30]]

            row = int(TIME_ROW_MAP.get(entry['time'])) + 8*self._date_to_row(entry['date']) 
            col = self._date_to_col(entry['date'])
            if col < 1 or row < 1:
                continue

            entry_text = f"{entry['student']}:{entry['subject']}"
            cell_primary = get_top_left_if_merged(new_ws, row, col)
            cell_secondary = get_top_left_if_merged(new_ws, row, col + 1)

            if not cell_primary.value:
                cell_primary.value = entry_text
            elif not cell_secondary.value:
                cell_secondary.value = entry_text
        os.makedirs(os.path.dirname(self.teacher_output_path), exist_ok=True)
        output_wb.save(self.teacher_output_path)
        
    def generate_student_excels(self):
        def normalize_grade(raw_grade):
            raw = raw_grade.lower()
            if any(k in raw for k in ['小', 'elementary', '小学']):
                return 'elementary'
            elif any(k in raw for k in ['中', 'middle', '中学']):
                return 'middle'
            elif any(k in raw for k in ['高', 'high', '高校']):
                return 'high'
            return None
        grouped_students = {'elementary': [], 'middle': [], 'high': []}
        for entry in self.schedule_data:
            
            grade_raw = entry.get('grade', '')
            grade = normalize_grade(grade_raw)
            if grade in grouped_students:
                grouped_students[grade].append(entry)

        for category, entries in grouped_students.items():
            if not entries:
                continue

            wb = Workbook()
            wb.remove(wb.active)
            students = sorted({
                _normalize_name(entry['student']) 
                for entry in entries 
                if entry.get('student')
            })
            
            for student in students:
                student_entries = [e for e in entries if _normalize_name(e['student']) == student]
         
                matched = False
                for student_templ in self.student_template:
                    if isinstance(student_templ, str):
                        template_wb = load_workbook(student_templ)
                    elif isinstance(student_templ, OpenpyxlWorkbook):
                        template_wb = student_templ
                    else:
                        raise TypeError("student_templ must be a path or openpyxl Workbook")

                    for sheet_name in template_wb.sheetnames:
                        student_norm = _normalize_name(student)
                        sheet_norm = _normalize_name(sheet_name)
                        # Direct full or partial match
                        if (
                            student_norm == sheet_norm or
                            student_norm in sheet_norm or
                            sheet_norm in student_norm or
                            any(part in sheet_norm for part in student_norm.split())
                        ):
    
                            template_ws = template_wb[sheet_name]
                            new_sheet_name = f"{sheet_name[:9]}"
                            new_ws = copy_worksheet_template(wb, template_ws, new_sheet_name)

                            for entry in student_entries:
                                
                                row = int(TIME_ROW_MAP.get(entry['time'])) + 8 * self._date_to_row(entry['date']) 
                                col = self._date_to_col(entry['date'])
                                if row and col:
                                    new_ws.cell(row=row, column=col-1).value = entry['subject']
                                    new_ws.cell(row=row, column=col).value = entry['type']
                                else:
                                    print(f"⚠️ Skipping invalid date/time in entry: {entry}")

                            matched = True
                            break  # Stop after first matching sheet

                if not matched:
                    print(f"⚠️ No matching sheet found for student: {student}")

            # Ensure workbook has at least one visible sheet
            if not wb.sheetnames:
                fallback_ws = wb.create_sheet("NoMatches")
                fallback_ws["A1"] = "⚠️ No matching sheets found for any student in this group."

            output_path = self.student_output_dirs[category]
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            print(f"💾 Saved: {output_path}")

    def _date_to_col(self, date_str):
        base_col = 2
        try:
            index = self.date_list.index(date_str) %5
        except ValueError:
            index = 0
        return base_col + index * 2 + 1

    def _date_to_row(self, date_str):
        try:
            index = int(self.date_list.index(date_str) /5)
        except ValueError:
            index = 0
        return index
