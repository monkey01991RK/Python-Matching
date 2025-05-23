import openpyxl
import re
from datetime import datetime
import json

class Teacher_data:
    def __init__(self, stu_file_path) -> None:
        self.stu_file_path = stu_file_path
        self.date_list = set() 
        self.teach_main()

    def _read_path(self, path) -> None:
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self.sheet_names = self.wb.sheetnames

    def _is_excluded(self, date_str, time_str):
        excluded_times = {
            "2025-04-05": [("13:10", "17:50")],
            "2025-04-12": [("13:10", "16:00")]
        }
        if date_str in excluded_times:
            for start, end in excluded_times[date_str]:
                if start <= time_str <= end:
                    return True
        return False

    def _has_diagonal(self, cell):
        if cell.border is None:
            return False
        return cell.border.diagonalUp or cell.border.diagonalDown

    def extract_schedule_calendar_blocks(self, file_path):
        results = {}
        self._read_path(file_path)

        for sheet_name in self.sheet_names:
            ws = self.wb[sheet_name]
            full_name = ""
            row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]

            # Extract teacher's name
            for i, cell in enumerate(row1):
                if cell.value and "講師名：" in str(cell.value):
                    next_cells = row1[i + 1:i + 4]
                    parts = [str(c.value).strip().replace("さん", "") for c in next_cells if c.value]
                    full_name = "".join(parts)
                    if not full_name:
                        full_name = "Unknown_Teacher"
                    break

            if not full_name:
                continue

            # Dynamically find calendar block start rows (e.g., rows with "3月" or "4月")
            calendar_starts = []
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val and re.search(r"[0-9０-９]{1,2}月", str(val)):
                    calendar_starts.append(row)

            for row_base in calendar_starts:
                month_cell = ws.cell(row=row_base, column=1).value
                if not month_cell:
                    continue

                month_match = re.search(r"([0-9０-９]{1,2})月", str(month_cell))
                if not month_match:
                    continue

                try:
                    month_str = month_match.group(1).translate(str.maketrans("０１２３４５６７８９", "0123456789"))
                    current_month = int(month_str)
                except:
                    continue

                for col in range(3, ws.max_column + 1, 2):
                    day_cell = ws.cell(row=row_base, column=col).value
                    if not day_cell or not re.search(r"\d{1,2}", str(day_cell)):
                        continue

                    try:
                        day = int(re.sub(r"[^\d]", "", str(day_cell)))
                        class_date = datetime(2025, current_month, day).date()
                        self.date_list.add(class_date)
                        date_str = class_date.isoformat()
                    except:
                        continue

                    for r_offset in range(2, 7):
                        time_cell = ws.cell(row=row_base + r_offset, column=1).value
                        if not time_cell:
                            continue

                        time_match = re.match(r"(\d{1,2})[:：](\d{2})", str(time_cell))
                        if not time_match:
                            continue

                        time_str = f"{time_match[1].zfill(2)}:{time_match[2]}"

                        if self._is_excluded(date_str, time_str):
                            continue

                        status = []
                        for offset in range(2):
                            check_cell = ws.cell(row=row_base + r_offset, column=col + offset)
                            is_free = not self._has_diagonal(check_cell)
                            status.append(is_free)
                            if not is_free:
                                # print(f"No diagonal: {check_cell.coordinate} - Value: {check_cell.value}")
                                continue

                            if full_name not in results:
                                results[full_name] = { 
                                    "t_sheetname":sheet_name,
                                    "schedule": {}
                                    }
                            if date_str not in results[full_name]["schedule"]:
                                results[full_name]["schedule"][date_str] = {}
                            results[full_name]["schedule"][date_str][time_str] = status

        return results
    def teach_main(self):
        all_schedules = self.extract_schedule_calendar_blocks(self.stu_file_path)
        output_path = "teacher_diagonal_schedule.json"
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(all_schedules, f, ensure_ascii=False, indent=2)
        print(f"Schedule exported to {output_path} with {sum(len(v['schedule']) for v in all_schedules.values())} dates total.")


        # Save date list
        sorted_dates = sorted(date.isoformat() for date in self.date_list)
        with open("lecture_dates.json", "w", encoding="utf-8") as f:
            json.dump(sorted_dates, f, ensure_ascii=False, indent=2)
