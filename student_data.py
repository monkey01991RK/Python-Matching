import pandas as pd
import json
from datetime import datetime
import re
import openpyxl


class Student_data:
    def __init__(self, stu_file_paths) -> None:
        self.stu_file_paths = stu_file_paths  # List of Excel file paths
        self.stu_main()

    def _read_path(self, path) -> None:
        self.xls = pd.ExcelFile(path)
        self.sheet_names = self.xls.sheet_names
        self.wb = openpyxl.load_workbook(path)
        
    def extract_schedule_calendar_blocks(self, file_path):
        results = {}
        self._read_path(file_path)

        for sheet in self.sheet_names:
            df = self.xls.parse(sheet, header=None)
            ws = self.wb[sheet]
            s_sheetname = sheet

            full_name = ""

            row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
            for i, cell in enumerate(row1):
                if cell.value and "生徒名" in str(cell.value):
                    # Collect values from the next few cells to the right (G, H, I)
                    next_cells = row1[i + 1:i + 4]  # get next 3 cells
                    parts = [str(c.value).strip().replace("さん", "") for c in next_cells if c.value]
                    full_name = "".join(parts)
                    break

            for date_idx in range(2, 49, 8):  # iterate through date rows
                if date_idx >= len(df):
                    continue
                month_row = df.iloc[date_idx]
                day_row = df.iloc[date_idx]

                for col in range(1, len(day_row), 2):
                    month_cell = "" if pd.isna(month_row[0]) else str(month_row[0])
                    month_match = re.search(r"([0-9０-９]{1,2})月", month_cell)
                    if not month_match:
                        continue
                    month_str = month_match[1].translate(str.maketrans("０１２３４５６７８９", "0123456789"))
                    current_month = int(month_str)

                    raw_day = "" if pd.isna(day_row[col]) else str(day_row[col])
                    if "日" not in raw_day:
                        continue
                    try:
                        day = int(re.sub(r"[^\d]", "", raw_day))
                        class_date = datetime(2025, current_month, day).date()
                    except ValueError:
                        continue

                    for row_idx in range(date_idx + 2, date_idx + 7):
                        if row_idx >= len(df):
                            continue
                        time_val = df.iloc[row_idx, 0]
                        if pd.isna(time_val):
                            continue
                        time_match = re.match(r"\d{1,2}:\d{2}", str(time_val))
                        if not time_match:
                            continue
                        time_str = time_match[0]

                        col_letter1 = openpyxl.utils.get_column_letter(col + 1)
                        col_letter2 = openpyxl.utils.get_column_letter(col + 2)
                        cell1 = ws[f"{col_letter1}{row_idx + 1}"]
                        cell2 = ws[f"{col_letter2}{row_idx + 1}"]

                        diag1 = cell1.border and (cell1.border.diagonalUp or cell1.border.diagonalDown)
                        diag2 = cell2.border and (cell2.border.diagonalUp or cell2.border.diagonalDown)

                        # ONLY output when both are false (no diagonal)
                        if not diag1 and not diag2:
                           
                            date_str = class_date.isoformat()
                            # Initialize student record if not yet added
                            if full_name not in results:
                                results[full_name] = {
                                    "s_sheetname": s_sheetname,
                                    "schedule": {}
                                }

                            # Initialize date if not yet added
                            if date_str not in results[full_name]["schedule"]:
                                results[full_name]["schedule"][date_str] = {}
                            # Mark time slot as available
                            results[full_name]["schedule"][date_str][time_str] = True
        return results
    def stu_main(self):
        all_data = {}

        for path in self.stu_file_paths:
            student_data = self.extract_schedule_calendar_blocks(path)
            for student, info in student_data.items():
                if student not in all_data:
                    all_data[student] = info
                else:
                    # Merge schedules
                    for date, slots in info["schedule"].items():
                        if date not in all_data[student]["schedule"]:
                            all_data[student]["schedule"][date] = slots
                        else:
                            all_data[student]["schedule"][date].update(slots)
        with open("student_schedules.json", "w", encoding="utf-8") as f:
            json.dump(all_data, f, ensure_ascii=False, indent=2)

 
